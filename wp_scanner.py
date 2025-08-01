#!/usr/bin/env python3
"""
WordPress Site Information Scanner
Retrieves WordPress version, plugins, and themes information from a WordPress installation
Outputs results to an Excel file
"""

import os
import re
from pathlib import Path
import argparse
from datetime import datetime
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ Error: openpyxl is required. Install it with: pip install openpyxl")
    exit(1)

class WordPressScanner:
    def __init__(self, wp_path):
        self.wp_path = Path(wp_path)
        self.results = {
            'scan_date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'wordpress_path': str(self.wp_path),
            'wordpress_version': None,
            'plugins': {},
            'themes': {}
        }
    
    def find_wordpress_version(self):
        """Extract WordPress version from version.php file"""
        version_file = self.wp_path / 'wp-includes' / 'version.php'
        
        if not version_file.exists():
            print(f"❌ WordPress version file not found at {version_file}")
            return None
        
        try:
            with open(version_file, 'r') as f:
                content = f.read()
                match = re.search(r"\$wp_version\s*=\s*'([^']+)'", content)
                if match:
                    self.results['wordpress_version'] = match.group(1)
                    print(f"✅ WordPress Version: {match.group(1)}")
                    return match.group(1)
        except Exception as e:
            print(f"❌ Error reading version file: {e}")
        
        return None
    
    def get_plugin_info(self, plugin_dir):
        """Extract plugin information from plugin headers"""
        plugin_info = {
            'name': plugin_dir.name,
            'version': 'Unknown',
            'description': '',
            'author': '',
            'plugin_uri': '',
            'directory': plugin_dir.name
        }
        
        # Find the main plugin file (usually matches directory name or contains plugin headers)
        php_files = list(plugin_dir.glob('*.php'))
        
        for php_file in php_files:
            try:
                with open(php_file, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read(8192)  # Read first 8KB (WordPress standard)
                    
                    # Check if this file contains plugin headers
                    if 'Plugin Name:' in content:
                        # Extract plugin information
                        patterns = {
                            'name': r'Plugin Name:\s*(.+)',
                            'version': r'Version:\s*(.+)',
                            'description': r'Description:\s*(.+)',
                            'author': r'Author:\s*(.+)',
                            'plugin_uri': r'Plugin URI:\s*(.+)'
                        }
                        
                        for key, pattern in patterns.items():
                            match = re.search(pattern, content, re.IGNORECASE)
                            if match:
                                plugin_info[key] = match.group(1).strip()
                        
                        # If we found plugin headers, we can stop searching
                        break
                        
            except Exception as e:
                print(f"  ⚠️  Error reading {php_file}: {e}")
        
        return plugin_info
    
    def scan_plugins(self):
        """Scan all plugins in wp-content/plugins directory"""
        plugins_dir = self.wp_path / 'wp-content' / 'plugins'
        
        if not plugins_dir.exists():
            print(f"❌ Plugins directory not found at {plugins_dir}")
            return
        
        print("\n🔍 Scanning Plugins...")
        plugin_dirs = [d for d in plugins_dir.iterdir() if d.is_dir() and not d.name.startswith('.')]
        
        for plugin_dir in sorted(plugin_dirs):
            plugin_info = self.get_plugin_info(plugin_dir)
            self.results['plugins'][plugin_dir.name] = plugin_info
            
            print(f"  📦 Found: {plugin_info['name']} (v{plugin_info['version']})")
    
    def get_theme_info(self, theme_dir):
        """Extract theme information from style.css"""
        theme_info = {
            'name': theme_dir.name,
            'version': 'Unknown',
            'description': '',
            'author': '',
            'theme_uri': '',
            'template': '',  # Parent theme if this is a child theme
            'directory': theme_dir.name
        }
        
        style_css = theme_dir / 'style.css'
        
        if style_css.exists():
            try:
                with open(style_css, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read(8192)  # Read first 8KB
                    
                    patterns = {
                        'name': r'Theme Name:\s*(.+)',
                        'version': r'Version:\s*(.+)',
                        'description': r'Description:\s*(.+)',
                        'author': r'Author:\s*(.+)',
                        'theme_uri': r'Theme URI:\s*(.+)',
                        'template': r'Template:\s*(.+)'
                    }
                    
                    for key, pattern in patterns.items():
                        match = re.search(pattern, content, re.IGNORECASE)
                        if match:
                            theme_info[key] = match.group(1).strip()
                            
            except Exception as e:
                print(f"  ⚠️  Error reading {style_css}: {e}")
        
        return theme_info
    
    def scan_themes(self):
        """Scan all themes in wp-content/themes directory"""
        themes_dir = self.wp_path / 'wp-content' / 'themes'
        
        if not themes_dir.exists():
            print(f"❌ Themes directory not found at {themes_dir}")
            return
        
        print("\n🎨 Scanning Themes...")
        theme_dirs = [d for d in themes_dir.iterdir() if d.is_dir() and not d.name.startswith('.')]
        
        for theme_dir in sorted(theme_dirs):
            theme_info = self.get_theme_info(theme_dir)
            self.results['themes'][theme_dir.name] = theme_info
            
            print(f"  🎭 Found: {theme_info['name']} (v{theme_info['version']})")
    
    def create_excel_report(self, output_file):
        """Create an Excel file with the scan results"""
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        info_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Create Overview sheet
        ws_overview = wb.active
        ws_overview.title = "Overview"
        
        # Add overview information
        overview_data = [
            ["WordPress Site Information Report", ""],
            ["", ""],
            ["Scan Date:", self.results['scan_date']],
            ["WordPress Path:", self.results['wordpress_path']],
            ["WordPress Version:", self.results['wordpress_version'] or "Unknown"],
            ["", ""],
            ["Summary", ""],
            ["Total Plugins:", len(self.results['plugins'])],
            ["Total Themes:", len(self.results['themes'])]
        ]
        
        for row_idx, row_data in enumerate(overview_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_overview.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=16)
                elif row_idx == 7:
                    cell.font = Font(bold=True, size=12)
                elif col_idx == 1 and row_idx > 2:
                    cell.font = Font(bold=True)
        
        ws_overview.column_dimensions['A'].width = 20
        ws_overview.column_dimensions['B'].width = 50
        
        # Create Plugins sheet
        ws_plugins = wb.create_sheet("Plugins")
        
        # Add headers
        plugin_headers = ["Plugin Name", "Version", "Author", "Description", "Plugin URI", "Directory"]
        for col_idx, header in enumerate(plugin_headers, 1):
            cell = ws_plugins.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add plugin data
        row_idx = 2
        for plugin_dir, plugin_info in sorted(self.results['plugins'].items()):
            ws_plugins.cell(row=row_idx, column=1, value=plugin_info['name']).border = border
            ws_plugins.cell(row=row_idx, column=2, value=plugin_info['version']).border = border
            ws_plugins.cell(row=row_idx, column=3, value=plugin_info['author']).border = border
            ws_plugins.cell(row=row_idx, column=4, value=plugin_info['description'][:255]).border = border  # Excel cell limit
            ws_plugins.cell(row=row_idx, column=5, value=plugin_info['plugin_uri']).border = border
            ws_plugins.cell(row=row_idx, column=6, value=plugin_info['directory']).border = border
            row_idx += 1
        
        # Auto-adjust column widths
        for column in ws_plugins.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_plugins.column_dimensions[column_letter].width = adjusted_width
        
        # Create Themes sheet
        ws_themes = wb.create_sheet("Themes")
        
        # Add headers
        theme_headers = ["Theme Name", "Version", "Author", "Description", "Theme URI", "Parent Theme", "Directory"]
        for col_idx, header in enumerate(theme_headers, 1):
            cell = ws_themes.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Add theme data
        row_idx = 2
        for theme_dir, theme_info in sorted(self.results['themes'].items()):
            ws_themes.cell(row=row_idx, column=1, value=theme_info['name']).border = border
            ws_themes.cell(row=row_idx, column=2, value=theme_info['version']).border = border
            ws_themes.cell(row=row_idx, column=3, value=theme_info['author']).border = border
            ws_themes.cell(row=row_idx, column=4, value=theme_info['description'][:255]).border = border
            ws_themes.cell(row=row_idx, column=5, value=theme_info['theme_uri']).border = border
            ws_themes.cell(row=row_idx, column=6, value=theme_info['template']).border = border
            ws_themes.cell(row=row_idx, column=7, value=theme_info['directory']).border = border
            row_idx += 1
        
        # Auto-adjust column widths
        for column in ws_themes.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_themes.column_dimensions[column_letter].width = adjusted_width
        
        # Save the file
        wb.save(output_file)
        print(f"\n💾 Excel report saved to: {output_file}")
    
    def print_summary(self):
        """Print a summary of the scan results"""
        print("\n" + "="*60)
        print("📊 SCAN SUMMARY")
        print("="*60)
        print(f"WordPress Version: {self.results['wordpress_version'] or 'Unknown'}")
        print(f"Total Plugins: {len(self.results['plugins'])}")
        print(f"Total Themes: {len(self.results['themes'])}")
        print("="*60)
    
    def run(self):
        """Run the complete scan"""
        print(f"🚀 Starting WordPress scan at: {self.wp_path}")
        print("="*60)
        
        # Check if path exists
        if not self.wp_path.exists():
            print(f"❌ Error: Path {self.wp_path} does not exist!")
            return False
        
        # Verify it's a WordPress installation
        wp_config = self.wp_path / 'wp-config.php'
        if not wp_config.exists():
            print(f"❌ Error: No wp-config.php found. Is this a WordPress installation?")
            return False
        
        # Run scans
        self.find_wordpress_version()
        self.scan_plugins()
        self.scan_themes()
        self.print_summary()
        
        return True


def main():
    parser = argparse.ArgumentParser(
        description='Scan WordPress installation for version, plugins, and themes information'
    )
    parser.add_argument(
        'wp_path',
        nargs='?',
        default='.',
        help='Path to WordPress installation (default: current directory)'
    )
    parser.add_argument(
        '-o', '--output',
        help='Output Excel file (default: wordpress_site_info.xlsx)',
        default='wordpress_site_info.xlsx'
    )
    
    args = parser.parse_args()
    
    # Convert to absolute path
    wp_path = Path(args.wp_path).resolve()
    
    # Create scanner and run
    scanner = WordPressScanner(wp_path)
    
    if scanner.run():
        scanner.create_excel_report(args.output)
        print("\n✅ Scan completed successfully!")
    else:
        print("\n❌ Scan failed!")
        exit(1)


if __name__ == "__main__":
    main()
